from openpyxl import Workbook
import bank_requester
import os


files_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'files')  


def create_workbook(success, errors, number):
    wb = Workbook(write_only=True)
    errors_ws = wb.create_sheet('Ошибки')
    for row in errors:
        errors_ws.append(row)
    success_ws = wb.create_sheet('Успешно')
    for row in success:
        success_ws.append(row)
    filename = os.path.join(files_path, f'Результат-{number}.xlsx')
    wb.save(filename)
    return filename