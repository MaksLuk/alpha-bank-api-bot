import datetime
import asyncio
import aiohttp
from openpyxl import load_workbook
import my_settings
import json
from database import Database
import os
import excel


headers = {
    "Content-Type": "application/json;charset=UTF-8",
    "API-key": my_settings.bank_api_key,
}


with open(os.path.join(excel.files_path, "fias-regions.json"), 'r', encoding='utf-8') as f:
    raw_json_data = f.read()
cities = json.loads(raw_json_data)


def time_check():
    ''' время в питоне можно сравнивать просто в формате "07:00:00" > "06:00:00" '''
    now_time = datetime.datetime.now().strftime("%H:%M:%S")
    return my_settings.start_working_time < now_time < my_settings.end_working_time


async def request(session, data):
    url = "https://partner.alfabank.ru/public-api/v2/leads"
    while not time_check:
        await asyncio.sleep(3*60)
    response = await session.post(url, headers=headers, json=data)
    if response.status == 200:
        return True, 200
    return False, response.status


async def parse_file(filename, now_count):
    ''' преобразует excel-файл в список словарей для отправки запросов '''
    success = [['ИНН', 'ФИО', 'Телефон', 'Организация', 'Сценарий', 'Город', 'Комментарий'], ]
    errors = [['ИНН', 'ФИО', 'Телефон', 'Организация', 'Сценарий', 'Город', 'Комментарий', 'Код ошибки'], ]
    wb = load_workbook(filename, read_only=True)
    sheet = wb.active
    async with aiohttp.ClientSession() as session:
        for i in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            i = [j.value for j in i]
            if not all(i[:4] + [i[5]]):
                errors.append(i + ['Указаны не все параметры'])
                await Database.insert_db((777, now_count) + tuple(i))
                continue

            city = cities.get(get_city(i[5]))
            if not city:
                errors.append(i + ['Город не найден'])
                await Database.insert_db((888, now_count) + tuple(i))
                continue
                
            i[2] = '+' + i[2] if i[2][0] == '7' else i[2]
            i[2] = '+7' + i[2][1:] if i[2][0] == '8' else i[2]

            response_bool, r_status = await request(session, create_payload(i, city))
            if response_bool:
                await Database.insert_db((1, now_count) + tuple(i))
            else:
                await Database.insert_db((r_status, now_count) + tuple(i))
            if not response_bool:
                errors.append(i + [f'Ошибка при запросе к API: {r_status}'])
                continue
            success.append(i)
            await asyncio.sleep(my_settings.timer)
    wb.close()
    return success, errors


def get_city(raw_city):
    ''' делает из "г. Москва и Московская область" просто "Москва" '''
    city = raw_city.strip().replace('г. ', '')
    if ' и ' in city:
        city = city[:city.find(' и ')]
    return city


def create_payload(i, city):
    info = {
        "organizationInfo": {
            "organizationName": i[3],
            "inn": i[0]
        },
        "contactInfo": [{
            "fullName": i[1],
            "phoneNumber": i[2],
        }],
        "requestInfo": { "cityCode": city },
        "productInfo": [{ "productCode": "LP_RKO" }],
    }     
    if i[6]:
        info["requestInfo"]["comment"] = i[6]
    if i[4] and i[4].strip() == "ДБ":
        info["requestInfo"]["advCode"] = "Действующий бизнес"
    return info

